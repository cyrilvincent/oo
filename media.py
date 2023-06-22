import datetime
from dataclasses import dataclass
from typing import List
import csv
import pickle
import xml.dom.minidom as dom
import jsonpickle
import openpyxl


@dataclass
class Author:

    first_name: str
    last_name: str

@dataclass
class Publisher:

    id: str
    name: str

class Media:

# Créer le constructeur, les attributs
# net_price = price * 1.055
# tester

# Transformer net_price en property
# Créer la dataclass Author
# Créer la dataclass Publisher
# Tester
    nb_media = 0

    def __init__(self, id: str,
                 title: str,
                 price: float,
                 date: datetime.datetime = datetime.datetime.now(),
                 publisher: Publisher = Publisher("", ""),
                 authors: List[Author] = []):
        self.id = id
        self.title = title
        self.price = price
        self.date = date
        self.publisher = publisher
        self.authors = authors
        Media.nb_media += 1

    @property
    def net_price(self):
        return self.price * 1.2

    def __del__(self):
        Media.nb_media -= 1

    def __repr__(self):
        return f"Media: {self.id} {self.title} {self.price}"

class Book(Media):

    nb_book = 0

    def __init__(self, id: str,
                 title: str,
                 price: float,
                 date: datetime.datetime = datetime.datetime.now(),
                 publisher: Publisher = Publisher("", ""),
                 authors: List[Author] = [],
                 nb_page = 0
                 ):
        super().__init__(id, title,price,date,publisher,authors)
        self.nb_page = nb_page
        Book.nb_book += 1

    @property
    def net_price(self):
        return self.price * 1.055 * 0.95 + 0.01

    def __del__(self):
        super().__del__()
        Book.nb_book -= 1

class Cd(Media):

    def __init__(self, id: str,
                 title: str,
                 price: float,
                 date: datetime.datetime = datetime.datetime.now(),
                 publisher: Publisher = Publisher("", ""),
                 authors: List[Author] = [],
                 nb_track = 0
                 ):
        super().__init__(id, title,price,date,publisher,authors)
        self.nb_track = nb_track

    @property
    def net_price(self):
        return self.price * 1.2 * 0.5

class Cart:

    def __init__(self):
        self.medias:List[Media] = []
        self.is_validate = False

    def add(self, media: Media):
        self.medias.append(media)

    def remove(self, media: Media):
        self.medias.remove(media)

    @property
    def count(self):
        return len(self.medias)

    @property
    def total_net_price(self) -> float:
        total = 0
        for m in self.medias:
            total += m.net_price
        return total

class MediaService:

    def __init__(self):
        self.medias: List[Media] = []

    def load_json(self, path: str):
        with open(path, "r") as f:
            s = f.read()
            self.medias = jsonpickle.decode(s)

    def clear(self):
        self.medias = []

    def load_xml(self, path: str):
        document: dom.Document = dom.parse(path)
        books = document.getElementsByTagName("book")
        for book in books:
            id = book.getAttribute("id")
            title = book.getAttribute("title")
            price = float(book.getAttribute("price"))
            b = Book(id, title, price)
            self.medias.append(b)

    def save_xml(self, path: str):
        document = dom.Document()
        books = document.createElement("books")
        document.appendChild(books)
        for m in self.medias:
            book = document.createElement("book")
            book.setAttribute("id", m.id)
            book.setAttribute("title", m.title)
            book.setAttribute("price", str(m.price))
            books.appendChild(book)
        with open(path, "w") as f:
            document.writexml(f, addindent="\t", newl="\n", )

    def load_xl(self, path: str):
        wb = openpyxl.open(path)
        sheet = wb.worksheets[0]
        for i in range(2, sheet.max_row + 1):
            id = sheet.cell(i, 1).value
            title = sheet.cell(i, 2).value
            price = sheet.cell(i, 3).value
            b = Book(id, title, price)
            self.medias.append(b)


# b1 = Book("001", "Python", 10)
# print(b1)
# print(f"Book price: {b1.net_price}")
# b2 = Book("002", "Numpy", 20, publisher=Publisher("123", "ENI"))
# b2.price+=1
# print(b2.price)
# print(b2)
#
# print(b1.net_price)
# # <=>
# # print(Book.net_price(b1))
#
# a1 = Author("Victor", "Hugo")
# print(a1)
#
# b3 = Book("1234", "Les misérables", 5, authors=[])
# b3.authors.append(Author("toto", "titi"))
#
# print(f"Nb media: {Media.nb_media}")
# del b3
# print(f"Nb media: {Media.nb_media}")
#
# cd1 = Cd("0014", "Johnny", 10, nb_track=9)
# print(cd1)






